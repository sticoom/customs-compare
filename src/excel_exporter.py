"""
Excel 报告导出
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.comparator import STATUS_PASS, STATUS_FAIL, STATUS_FUZZY, STATUS_MANUAL


# 样式定义
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")

PASS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
FUZZY_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
MANUAL_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_LABELS = {
    STATUS_PASS: "✅ 通过",
    STATUS_FAIL: "❌ 不通过",
    STATUS_FUZZY: "⚠️ 模糊匹配",
    STATUS_MANUAL: "🔍 待人工确认",
}


def get_fill(status: str) -> PatternFill:
    return {
        STATUS_PASS: PASS_FILL,
        STATUS_FAIL: FAIL_FILL,
        STATUS_FUZZY: FUZZY_FILL,
        STATUS_MANUAL: MANUAL_FILL,
    }.get(status, PatternFill())


def export_to_excel(comparison_result: dict, filename: str) -> str:
    """
    将比对结果导出为 Excel 文件
    返回: 保存的文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "比对结果"

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = f"报关单 vs 预录单 比对报告 - 合同协议号: {comparison_result.get('contract_no', '')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 概览
    summary = comparison_result["summary"]
    ws.merge_cells("A2:F2")
    ws["A2"] = f"✅ 通过: {summary['pass_count']}  |  ❌ 不通过: {summary['fail_count']}  |  ⚠️ 模糊: {summary['fuzzy_count']}  |  🔍 待确认: {summary['manual_count']}"
    ws["A2"].font = Font(size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4

    # ---- 第一部分：表头字段 ----
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "一、表头字段（项号前）"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    # 表头
    headers = ["字段名称", "报关单值", "预录单值", "比对状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    for result in comparison_result["header_results"]:
        ws.cell(row=row, column=1, value=result["label"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=result["customs_value"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=result["pre_value"]).border = THIN_BORDER
        status_cell = ws.cell(row=row, column=4, value=STATUS_LABELS.get(result["status"], ""))
        status_cell.border = THIN_BORDER
        status_cell.fill = get_fill(result["status"])
        ws.cell(row=row, column=5, value=result.get("notes", "")).border = THIN_BORDER
        row += 1

    row += 1

    # ---- 第二部分：问题项汇总 ----
    problem_rows = []
    for r in comparison_result["header_results"]:
        if r["status"] in (STATUS_FAIL, STATUS_FUZZY):
            problem_rows.append(("表头", "-", r["label"], r["customs_value"], r["pre_value"], r["status"], r.get("notes", "")))
    for item_result in comparison_result["item_results"]:
        for f in item_result["fields"]:
            if f["status"] in (STATUS_FAIL, STATUS_FUZZY):
                problem_rows.append(("商品明细", item_result["item_no"], f["label"], f["customs_value"], f["pre_value"], f["status"], f.get("notes", "")))

    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "二、问题项汇总（不通过 + 模糊）"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    problem_headers = ["位置", "项号", "字段名称", "报关单值", "预录单值", "比对状态", "备注"]
    for col, header in enumerate(problem_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    for prow in problem_rows:
        for col, val in enumerate(prow, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 6:
                cell.value = STATUS_LABELS.get(val, "")
                cell.fill = get_fill(val)
        row += 1

    row += 1

    # ---- 第三部分：商品明细 ----
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "三、商品明细（项号后）"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    for item_result in comparison_result["item_results"]:
        # 项号标题
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"项号 {item_result['item_no']}"
        ws[f"A{row}"].font = Font(bold=True, size=11)
        row += 1

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        row += 1

        for field in item_result["fields"]:
            ws.cell(row=row, column=1, value=field["label"]).border = THIN_BORDER
            ws.cell(row=row, column=2, value=field["customs_value"]).border = THIN_BORDER
            ws.cell(row=row, column=3, value=field["pre_value"]).border = THIN_BORDER
            status_cell = ws.cell(row=row, column=4, value=STATUS_LABELS.get(field["status"], ""))
            status_cell.border = THIN_BORDER
            status_cell.fill = get_fill(field["status"])
            ws.cell(row=row, column=5, value=field.get("notes", "")).border = THIN_BORDER
            row += 1

        row += 1

    # 设置列宽
    col_widths = [10, 8, 20, 35, 35, 15, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filename)
    return filename


def export_multiple_to_excel(comparison_results: list, filename: str) -> str:
    """
    将多组比对结果导出到一个 Excel 文件，每组一个 sheet
    """
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for idx, result in enumerate(comparison_results):
        contract_no = result.get("contract_no", f"未知{idx + 1}")
        sheet_title = contract_no[:28] if len(contract_no) > 28 else contract_no
        ws = wb.create_sheet(title=sheet_title)

        # 标题
        ws.merge_cells("A1:F1")
        ws["A1"] = f"报关单 vs 预录单 比对报告 - 合同协议号: {contract_no}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # 文件来源
        c_files = result.get("customs_filenames", [])
        p_files = result.get("pre_filenames", [])
        ws.merge_cells("A2:F2")
        ws["A2"] = f"报关单: {', '.join(c_files)}  |  预录单: {', '.join(p_files)}"
        ws["A2"].font = Font(size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

        # 概览
        summary = result["summary"]
        ws.merge_cells("A3:F3")
        ws["A3"] = f"✅ 通过: {summary['pass_count']}  |  ❌ 不通过: {summary['fail_count']}  |  ⚠️ 模糊: {summary['fuzzy_count']}  |  🔍 待确认: {summary['manual_count']}"
        ws["A3"].font = Font(size=11)
        ws["A3"].alignment = Alignment(horizontal="center")

        row = 5

        # ---- 第一部分：表头字段 ----
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "一、表头字段（项号前）"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        headers = ["字段名称", "报关单值", "预录单值", "比对状态", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        row += 1

        for r in result["header_results"]:
            ws.cell(row=row, column=1, value=r["label"]).border = THIN_BORDER
            ws.cell(row=row, column=2, value=r["customs_value"]).border = THIN_BORDER
            ws.cell(row=row, column=3, value=r["pre_value"]).border = THIN_BORDER
            status_cell = ws.cell(row=row, column=4, value=STATUS_LABELS.get(r["status"], ""))
            status_cell.border = THIN_BORDER
            status_cell.fill = get_fill(r["status"])
            ws.cell(row=row, column=5, value=r.get("notes", "")).border = THIN_BORDER
            row += 1

        row += 1

        # ---- 第二部分：问题项汇总 ----
        problem_rows = []
        for r in result["header_results"]:
            if r["status"] in (STATUS_FAIL, STATUS_FUZZY):
                problem_rows.append(("表头", "-", r["label"], r["customs_value"], r["pre_value"], r["status"], r.get("notes", "")))
        for item_result in result["item_results"]:
            for f in item_result["fields"]:
                if f["status"] in (STATUS_FAIL, STATUS_FUZZY):
                    problem_rows.append(("商品明细", item_result["item_no"], f["label"], f["customs_value"], f["pre_value"], f["status"], f.get("notes", "")))

        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "二、问题项汇总（不通过 + 模糊）"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        problem_headers = ["位置", "项号", "字段名称", "报关单值", "预录单值", "比对状态", "备注"]
        for col, header in enumerate(problem_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        row += 1

        for prow in problem_rows:
            for col, val in enumerate(prow, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                if col == 6:
                    cell.value = STATUS_LABELS.get(val, "")
                    cell.fill = get_fill(val)
            row += 1

        row += 1

        # ---- 第三部分：商品明细 ----
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "三、商品明细（项号后）"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        for item_result in result["item_results"]:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"项号 {item_result['item_no']}"
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 1

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = HEADER_FONT_WHITE
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER
            row += 1

            for field in item_result["fields"]:
                ws.cell(row=row, column=1, value=field["label"]).border = THIN_BORDER
                ws.cell(row=row, column=2, value=field["customs_value"]).border = THIN_BORDER
                ws.cell(row=row, column=3, value=field["pre_value"]).border = THIN_BORDER
                status_cell = ws.cell(row=row, column=4, value=STATUS_LABELS.get(field["status"], ""))
                status_cell.border = THIN_BORDER
                status_cell.fill = get_fill(field["status"])
                ws.cell(row=row, column=5, value=field.get("notes", "")).border = THIN_BORDER
                row += 1

            row += 1

        # 设置列宽
        col_widths = [10, 8, 20, 35, 35, 15, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filename)
    return filename
